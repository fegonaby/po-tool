"""Base Excel export using openpyxl templates.

Loads a template workbook, writes rows starting at row 13 (header at row 12),
and returns the file as in-memory bytes.
"""

import io
from pathlib import Path
from typing import Sequence

import openpyxl


TEMPLATES_DIR = Path(__file__).resolve().parent.parent.parent / "templates"

HEADER_ROW = 12
DATA_START_ROW = 13


def build_excel(
    template_name: str,
    rows: list[dict],
    column_mapping: list[tuple[str, str]],
    *,
    sheet_name: str = "Sheet 1",
) -> bytes:
    """Populate an Excel template with data and return as bytes.

    Args:
        template_name: Filename within the templates/ directory (e.g. "DPO.xlsx").
        rows: List of result dicts from Databricks.
        column_mapping: Ordered list of (dict_key, header_label) pairs.
            Column order in the Excel follows this list.
        sheet_name: Name of the sheet to populate.

    Returns:
        In-memory Excel file bytes.
    """
    template_path = TEMPLATES_DIR / template_name
    wb = openpyxl.load_workbook(template_path)
    ws = wb[sheet_name]

    for col_idx, (_key, label) in enumerate(column_mapping, start=1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=label)

    for row_idx, row_data in enumerate(rows, start=DATA_START_ROW):
        for col_idx, (key, _label) in enumerate(column_mapping, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_excel_to_file(data: bytes, output_path: str | Path) -> Path:
    """Write Excel bytes to disk. Returns the Path written."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)
    return path
